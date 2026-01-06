import pandas as pd
import datetime
import re
import logging
import sys
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurar logging para ver errores
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('transformacion_errores.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

# Diccionario de mapeo para desarrollos
mapeo_desarrollos = {
    'Ciudad Deportiva': {
        'equivalencia_base': 'CD SUBC',
        'desarrollo_maestro': 'CIUDAD DEPORTIVA',
        'formato': 'base_espacio_numero'
    },
    'Demo': {
        'equivalencia_base': 'DEMO',
        'desarrollo_maestro': 'DEMO',
        'formato': 'solo_base'
    },
    'Fundadores': {
        'equivalencia_base': 'F',
        'desarrollo_maestro': 'FUNDADORES',
        'formato': 'solo_base'
    },
    'Hunucma': {
        'equivalencia_base': 'H',
        'desarrollo_maestro': 'HUNUCMA',
        'formato': 'solo_base'
    },
    'Punta Helena': {
        'equivalencia_base': 'PH',
        'desarrollo_maestro': 'PUNTA HELENA',
        'formato': 'base_numero'
    },
    'Parque Pimienta': {
        'equivalencia_base': 'A',
        'desarrollo_maestro': 'COIN',
        'formato': 'base_numero'
    },
    'Cumbres De La Hacienda': {
        'equivalencia_base': 'CUMBRES SUBC',
        'desarrollo_maestro': 'CUMBRES',
        'formato': 'base_espacio_numero'
    },
    'Telchac': {
        'equivalencia_base': 'T',
        'desarrollo_maestro': 'TELCHAC',
        'formato': 'base_numero'
    },
    'Terramarket': {
        'equivalencia_base': 'TM',
        'desarrollo_maestro': 'TERRAMARKET',
        'formato': 'solo_base'
    },
    'San Roque': {
        'equivalencia_base': 'SR RESID',
        'desarrollo_maestro': 'SAN ROQUE',
        'formato': 'base_espacio_numero'
    },
    'Puerto Telchac': {
        'equivalencia_base': 'PT SUBC',
        'desarrollo_maestro': 'PUERTO TELCHAC',
        'formato': 'base_espacio_numero'
    },
    'Mareta': {
        'equivalencia_base': 'PT MARETA',
        'desarrollo_maestro': 'PUERTO TELCHAC',
        'formato': 'solo_base'
    },
    'San Eduardo': {
        'equivalencia_base': {
            'P': 'SE RESID',
            'C': 'SE COMER'
        },
        'desarrollo_maestro': 'SAN EDUARDO',
        'formato': 'base_espacio_numero'
    },
    'Business Center': {
        'equivalencia_base': {
            'MF': 'MF',
            'F&F': 'F&F'
        },
        'desarrollo_maestro': 'COIN BUSINESS CENTER',
        'formato': 'solo_base'
    },
    'Santa Clara': {
        'equivalencia_base': 'SC SUBC',
        'desarrollo_maestro': 'SANTA CLARA',
        'formato': 'base_espacio_numero'
    },
    'Bosques De La Hacienda': {
        'equivalencia_base': 'SUBC',
        'desarrollo_maestro': 'HACIENDA TVA',
        'formato': 'base_espacio_numero'
    },
    'Jardines De La Hacienda': {
        'equivalencia_base': 'SUBC',
        'desarrollo_maestro': 'HACIENDA TVA',
        'formato': 'base_espacio_numero'
    },
    'Paseo Flamboyanes': {
        'equivalencia_base': 'SUBC',
        'desarrollo_maestro': 'HACIENDA TVA',
        'formato': 'base_espacio_numero'
    },
    'Paseo Henequen': {
        'equivalencia_base': 'SUBC',
        'desarrollo_maestro': 'HACIENDA TVA',
        'formato': 'base_espacio_numero'
    },
    'Paseo Ceiba': {
        'equivalencia_base': 'SUBC',
        'desarrollo_maestro': 'HACIENDA TVA',
        'formato': 'base_espacio_numero'
    },
    'Prolongacion': {
        'equivalencia_base': 'HDA PROL',
        'desarrollo_maestro': 'HACIENDA TVA',
        'formato': 'solo_base'
    },
    'Custo': {
        'equivalencia_base': 'CUSTO',
        'desarrollo_maestro': 'CUSTO',
        'formato': 'solo_base'
    },
    'Playaviva Apartments': {
        'equivalencia_base': None,  # Caso especial - se maneja en la lógica
        'desarrollo_maestro': 'COIN APARTMENTS',
        'formato': 'especial_playaviva'
    }
}

# Mapeo especial para etapas no numéricas
mapeo_etapas_especiales = {
    'Puerto Telchac': {
        'Carey': "9",
        'Arena': '19',
        'Coral': '20', 
        'Arrecife': '22'
    }
}

def extraer_tipo_y_numero_etapa(etapa, desarrollo):
    """Extrae el tipo (P, C, etc.) y número de la etapa"""
    try:
        if pd.isna(etapa) or etapa in ['N/A', '', ' ', None]:
            return None, None
        
        etapa_str = str(etapa).strip()
        
        # Primero verificar si hay un mapeo especial para esta combinación desarrollo-etapa
        if desarrollo in mapeo_etapas_especiales:
            if etapa_str in mapeo_etapas_especiales[desarrollo]:
                return None, mapeo_etapas_especiales[desarrollo][etapa_str]
        
        # Extraer tipo (letras) y número
        tipo_match = re.match(r'^([A-Za-z]+)', etapa_str)
        numero_match = re.search(r'\d+', etapa_str)
        
        tipo = tipo_match.group(1).upper() if tipo_match else None
        numero = numero_match.group() if numero_match else None
        
        return tipo, numero
        
    except Exception as e:
        logging.error(f"Error extrayendo tipo y número de etapa '{etapa}' para desarrollo '{desarrollo}': {str(e)}")
        logging.debug(traceback.format_exc())
        return None, None

def generar_equivalencia(etapa, desarrollo):
    """Genera la equivalencia según las reglas especificadas"""
    try:
        if desarrollo not in mapeo_desarrollos:
            logging.warning(f"Desarrollo no encontrado en mapeo: '{desarrollo}'")
            return None
        
        desarrollo_data = mapeo_desarrollos[desarrollo]
        formato = desarrollo_data.get('formato', 'base_espacio_numero')
        
        # Caso especial para Playaviva Apartments
        if desarrollo == 'Playaviva Apartments':
            if pd.isna(etapa) or etapa in ['N/A', '', ' ', None]:
                return 'N/A'
            else:
                # Para etapas no vacías, mantener el valor original de la etapa
                return str(etapa).strip()
        
        # Para los demás desarrollos, continuar con la lógica normal
        tipo_etapa, numero_etapa = extraer_tipo_y_numero_etapa(etapa, desarrollo)
        
        # Determinar la base de equivalencia según el tipo de etapa
        equivalencia_base = desarrollo_data['equivalencia_base']
        if isinstance(equivalencia_base, dict):
            # Si hay múltiples bases según el tipo de etapa
            if tipo_etapa in equivalencia_base:
                equivalencia_base = equivalencia_base[tipo_etapa]
            else:
                # Si no hay tipo específico, usar el primero disponible
                equivalencia_base = list(equivalencia_base.values())[0]
        
        if formato == 'solo_base':
            return equivalencia_base
        elif numero_etapa:
            if formato == 'base_espacio_numero':
                return f"{equivalencia_base} {numero_etapa}"
            elif formato == 'base_numero':
                return f"{equivalencia_base}{numero_etapa}"
        else:
            return equivalencia_base
            
    except Exception as e:
        logging.error(f"Error generando equivalencia para Etapa '{etapa}', Desarrollo '{desarrollo}': {str(e)}")
        logging.debug(traceback.format_exc())
        return None

def obtener_desarrollo_maestro(desarrollo):
    """Obtiene el desarrollo maestro del mapeo"""
    try:
        if desarrollo in mapeo_desarrollos:
            return mapeo_desarrollos[desarrollo]['desarrollo_maestro']
        else:
            logging.warning(f"Desarrollo maestro no encontrado para: '{desarrollo}'")
            return str(desarrollo).upper()
    except Exception as e:
        logging.error(f"Error obteniendo desarrollo maestro para '{desarrollo}': {str(e)}")
        logging.debug(traceback.format_exc())
        return str(desarrollo).upper() if desarrollo else "DESCONOCIDO"

def validar_fila(row, index):
    """Valida una fila individual y registra advertencias"""
    try:
        if pd.isna(row['Desarrollo']):
            logging.warning(f"Fila {index}: Desarrollo está vacio")
            return False
            
        desarrollo = str(row['Desarrollo']).strip()
        if desarrollo not in mapeo_desarrollos:
            logging.warning(f"Fila {index}: Desarrollo '{desarrollo}' no esta mapeado")
            
        return True
    except Exception as e:
        logging.error(f"Error validando fila {index}: {str(e)}")
        return False

def aplicar_formato_jesus_herrera(worksheet):
    """Aplica formato a la hoja JesusHerrera"""
    try:
        # Definir el estilo para los encabezados
        fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar formato a los encabezados
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border
            # Convertir a mayúsculas
            cell.value = str(cell.value).upper() if cell.value else ""
        
        # Ajustar el ancho de las columnas
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)
            
        logging.info("Formato aplicado a la hoja JesusHerrera")
        
    except Exception as e:
        logging.error(f"Error aplicando formato a JesusHerrera: {str(e)}")
        logging.debug(traceback.format_exc())

def procesar_excel(archivo_entrada, archivo_salida):
    """Función principal para procesar el CSV y generar Excel"""
    try:
        logging.info(f"Cargando archivo: {archivo_entrada}")
        
        # Leer CSV con manejo de diferentes codificaciones
        try:
            df = pd.read_csv(archivo_entrada)
        except UnicodeDecodeError:
            df = pd.read_csv(archivo_entrada, encoding='latin-1')
            
        logging.info(f"Archivo cargado. Filas: {len(df)}")
        
        # Validar columnas requeridas
        columnas_requeridas = ['Etapa', 'Desarrollo']
        for columna in columnas_requeridas:
            if columna not in df.columns:
                raise ValueError(f"Columna requerida no encontrada: {columna}")
        
        # Validar filas antes de procesar
        logging.info("Validando datos")
        for index, row in df.iterrows():
            validar_fila(row, index)
        
        # Aplicar las transformaciones
        logging.info("Aplicando transformaciones...")
        
        df['Equivalencia'] = df.apply(
            lambda row: generar_equivalencia(row['Etapa'], row['Desarrollo']), 
            axis=1
        )
        
        df['Desarrollo_Maestro'] = df['Desarrollo'].apply(obtener_desarrollo_maestro)
        
        # Validar resultados
        filas_sin_equivalencia = df['Equivalencia'].isna().sum()
        if filas_sin_equivalencia > 0:
            logging.warning(f"{filas_sin_equivalencia} filas sin equivalencia generada")
            
        # Mostrar resumen de transformaciones
        desarrollos_procesados = df['Desarrollo'].value_counts()
        logging.info("Resumen por desarrollo:")
        for desarrollo, count in desarrollos_procesados.items():
            logging.info(f"  {desarrollo}: {count} filas")
        
        # Crear DataFrame para la hoja BI (mantener todas las columnas originales más las nuevas)
        df_bi = df.copy()
        
        # Asegurarse de que las columnas estén en el orden correcto para BI
        columnas_bi = [
            'id_venta', 'Marca', 'Etapa', 'Equivalencia', 'Desarrollo_Maestro', 
            'Desarrollo', 'combinado', 'Etapa2', 'Unidad', 'Tipo', 'M2_Accion', 
            'PrecioM2_Accion', 'PrecioVenta', 'Asesor', 'Zona', 'Int_Ext', 
            'Eq', 'Cliente', 'F_Venta', 'Enganche', 'Cobrado', 
            'PrecioVenta_Pagado__Saldo', 'Comision_DireccionCU', 'Estatus', 
            'Cobrado_Enganche'
        ]
        columnas_existentes = [col for col in columnas_bi if col in df_bi.columns]
        df_bi = df_bi[columnas_existentes]

        # Crear DataFrame para la hoja JesusHerrera
        df_jesus = pd.DataFrame()
        
        # Mapeo de columnas
        df_jesus['MARCA'] = df['Marca']
        df_jesus['ETAPA'] = df['Equivalencia']
        df_jesus['DESARROLLO'] = df['Desarrollo_Maestro']
        df_jesus['COMBINADO'] = df['Equivalencia'] + " " + df['Unidad'].astype(str)
        df_jesus['ETAPA2'] = df['Etapa2']
        df_jesus['UNIDAD'] = df['Unidad']
        df_jesus['TIPO'] = df['Tipo']
        df_jesus['M2 / ACCION'] = df['M2_Accion']
        df_jesus['$ M2 / ACCION'] = df['PrecioM2_Accion']
        df_jesus['$ VENTA'] = df['PrecioVenta']
        df_jesus['ASESOR'] = df['Asesor']
        df_jesus['ZONA'] = df['Zona']
        df_jesus['INT / EXT'] = df['Int_Ext']
        df_jesus['EQ'] = df['Eq']
        df_jesus['CLIENTE'] = df['Cliente']
        df_jesus['F. VENTA'] = df['F_Venta']
        df_jesus['ENGANCHE'] = df['Enganche']
        df_jesus['COBRADO'] = df['Cobrado']
        df_jesus['SALDO'] = df['PrecioVenta_Pagado__Saldo']
        df_jesus['COMISION DIRECCION (C/U)'] = df['Comision_DireccionCU']
        df_jesus['ESTATUS'] = df['Estatus']
        df_jesus['COBRADO - ENGANCHE'] = df['Cobrado_Enganche']

        # Crear el archivo Excel
        with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:

            df_bi.to_excel(writer, sheet_name='BI', index=False)
            df_jesus.to_excel(writer, sheet_name='JesusHerrera', index=False)
            
            # Obtener el workbook y aplicar formato a la hoja JesusHerrera
            workbook = writer.book
            worksheet_jesus = workbook['JesusHerrera']
            aplicar_formato_jesus_herrera(worksheet_jesus)
        
        logging.info(f"Transformación completada. Archivo guardado: {archivo_salida}")
        logging.info(f"Resumen: {len(df)} filas procesadas")
        logging.info(f"Hojas creadas: 'BI' ({len(df_bi)} filas) y 'JesusHerrera' ({len(df_jesus)} filas)")

        return True
        
    except FileNotFoundError:
        logging.error(f"Archivo no encontrado: {archivo_entrada}")
        return False
    except Exception as e:
        logging.error(f"Error procesando CSV: {str(e)}")
        logging.debug(traceback.format_exc())
        return False

# Función para probar casos específicos
def probar_casos():
    """Función para probar casos específicos y verificar resultados"""
    casos_prueba = [
        # Casos originales
        ('P22', 'Ciudad Deportiva', 'CD SUBC 22', 'CIUDAD DEPORTIVA'),
        ('N/A', 'Demo', 'DEMO', 'DEMO'),
        ('N/A', 'Fundadores', 'F', 'FUNDADORES'),
        ('N/A', 'Hunucma', 'H', 'HUNUCMA'),
        ('PH ET 1 BIS', 'Punta Helena', 'PH1', 'PUNTA HELENA'),
        ('P1', 'Cumbres De La Hacienda', 'CUMBRES SUBC 1', 'CUMBRES'),
        ('N/A', 'Punta Helena', 'PH', 'PUNTA HELENA'),
        ('A1', 'Parque Pimienta', 'A1', 'COIN'),
        ('P1', 'Telchac', 'T1', 'TELCHAC'),
        ('N/A', 'Terramarket', 'TM', 'TERRAMARKET'),
        ('P1', 'San Roque', 'SR RESID 1', 'SAN ROQUE'),
        ('Arena', 'Puerto Telchac', 'PT SUBC 19', 'PUERTO TELCHAC'),
        ('Coral', 'Puerto Telchac', 'PT SUBC 20', 'PUERTO TELCHAC'),
        ('Arrecife', 'Puerto Telchac', 'PT SUBC 22', 'PUERTO TELCHAC'),
        ('P1', 'San Eduardo', 'SE RESID 1', 'SAN EDUARDO'),
        
        # Casos intermedios
        ('C1', 'San Eduardo', 'SE COMER 1', 'SAN EDUARDO'),
        ('P1', 'Santa Clara', 'SC SUBC 1', 'SANTA CLARA'),
        ('P1', 'Bosques De La Hacienda', 'SUBC 1', 'HACIENDA TVA'),
        ('P2', 'Jardines De La Hacienda', 'SUBC 2', 'HACIENDA TVA'),
        ('P3', 'Paseo Flamboyanes', 'SUBC 3', 'HACIENDA TVA'),
        ('P4', 'Paseo Henequen', 'SUBC 4', 'HACIENDA TVA'),
        ('P5', 'Paseo Ceiba', 'SUBC 5', 'HACIENDA TVA'),
        ('N/A', 'Prolongacion', 'HDA PROL', 'HACIENDA TVA'),
        
        # Últimos casos especiales
        ('MF', 'Business Center', 'MF', 'COIN BUSINESS CENTER'),
        ('N/A', 'Mareta', 'PT MARETA', 'PUERTO TELCHAC'),
        ('LA', 'Custo', 'CUSTO', 'CUSTO'),
        ('N/A', 'Playaviva Apartments', 'N/A', 'COIN APARTMENTS'),
        ('F1', 'Playaviva Apartments', 'F1', 'COIN APARTMENTS')
    ]
    
    logging.info("Probando casos de prueba...")
    exitoso = True
    
    for i, (etapa, desarrollo, equivalencia_esperada, desarrollo_maestro_esperado) in enumerate(casos_prueba):
        equivalencia_calculada = generar_equivalencia(etapa, desarrollo)
        desarrollo_maestro_calculado = obtener_desarrollo_maestro(desarrollo)
        
        if equivalencia_calculada == equivalencia_esperada and desarrollo_maestro_calculado == desarrollo_maestro_esperado:
            logging.info(f"✓ Caso {i+1}: CORRECTO - {etapa} + {desarrollo} → {equivalencia_calculada} | {desarrollo_maestro_calculado}")
        else:
            logging.error(f"✗ Caso {i+1}: ERROR")
            logging.error(f"  Entrada: Etapa '{etapa}', Desarrollo '{desarrollo}'")
            logging.error(f"  Esperado: '{equivalencia_esperada}', '{desarrollo_maestro_esperado}'")
            logging.error(f"  Obtenido: '{equivalencia_calculada}', '{desarrollo_maestro_calculado}'")
            exitoso = False
    
    return exitoso

# Ejemplo de uso
if __name__ == "__main__":
    # Probar casos primero
    logging.info("=== INICIANDO PRUEBAS ===")
    pruebas_exitosas = probar_casos()
    
    if pruebas_exitosas:
        logging.info("=== TODAS LAS PRUEBAS PASARON ===")

        def obtener_mes_anterior():
            date = datetime.date.today()

            # Ir al primer día del mes actual y retroceder un día
            ultimo_dia_mes_anterior = date.replace(day=1) - datetime.timedelta(days = 1)
            return ultimo_dia_mes_anterior.month, ultimo_dia_mes_anterior.year

        mes, year = obtener_mes_anterior()
        mes_str = f"{mes:02d}"

        archivo_entrada = 'data/reporte-cierre-de-mes.csv'
        archivo_salida = f"data/reporte-cierre-de-mes-transformado-{mes_str}-{year}.xlsx"
        exito = procesar_excel(archivo_entrada, archivo_salida)
        
        if exito:
            print("Transformacion completada exitosamente.")
            print(f"Archivo Excel generado: {archivo_salida}")
            print("Revisa 'transformacion_errores.log' para ver cualquier advertencia.")
        else:
            print("Error en la transformacion. Revisa el log para más detalles.")
    else:
        logging.error("=== FALLA EN LAS PRUEBAS ===")
        print("Algunas pruebas fallaron. Revisa el log para más detalles.")
        print("No se procesara el archivo CSV hasta que todas las pruebas pasen.")